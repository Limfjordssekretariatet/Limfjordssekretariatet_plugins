import os
import math

from qgis.core import (
    QgsProcessingAlgorithm,
    QgsProcessingContext,
    QgsProcessingFeedback,
    QgsProcessingParameterVectorLayer,
    QgsProcessingOutputNumber,
    QgsProcessingException,
    QgsProject,
)


# Sørg for at scriptets egen mappe er på sys.path, så 'utils' altid kan
# importeres uanset hvordan QGIS' script-provider loader filen.
import os as _os, sys as _sys
try:
    _d = _os.path.dirname(_os.path.abspath(__file__))
    if _d and _d not in _sys.path:
        _sys.path.insert(0, _d)
except NameError:
    pass
from utils import find_resultater_mappe, get_resultat_excel_navn


def find_regneark():
    """Resultater/Resultat_<omraade>.xlsx, eller None hvis den ikke findes."""
    resultater = find_resultater_mappe()
    if not resultater:
        return None
    sti = os.path.join(resultater, get_resultat_excel_navn())
    return sti if os.path.isfile(sti) else None


def genberegn_i_excel(excel_sti, feedback):
    """Faa Excel til at regne arkets formler igennem og gemme.

    openpyxl skriver tal, men beregner ikke formler. Cellen med N-tabet er en
    formel, og dens gemte vaerdi staar derfor tom, indtil regnearket har
    vaeret aabnet i Excel. Er Excel paa maskinen, ordnes det her; ellers maa
    brugeren aabne og gemme filen én gang.

    Returnerer True hvis genberegningen lykkedes.
    """
    try:
        import win32com.client
    except ImportError:
        feedback.pushInfo(
            "Excel kunne ikke fjernstyres herfra (pywin32 mangler) — "
            "formlerne skal regnes ud ved at åbne regnearket.")
        return False
    excel = None
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        bog = excel.Workbooks.Open(os.path.abspath(excel_sti))
        excel.CalculateFullRebuild()
        bog.Save()
        bog.Close(SaveChanges=False)
        feedback.pushInfo("Regnearkets formler blev regnet igennem i Excel.")
        return True
    except Exception as e:
        feedback.pushWarning(f"Kunne ikke regne formlerne igennem i Excel: {e}")
        return False
    finally:
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass


EXCEL_ARK_TILFOERSEL  = "1. Tilførsel"
EXCEL_ARK_OMSAETNING  = "2. Omsætning"
TN_BELASTNING_CELLE   = "C30"   # kg N/ha/år  — Tilførsel, vandoplandet
OMSAETNINGSRATE_CELLE = "D32"   # kg N/ha/døgn — Omsætning, Oversvømmelse


class BeregnOmsaetningsrate(QgsProcessingAlgorithm):

    PROJEKTOMRAADE = "PROJEKTOMRAADE"
    OUTPUT_RATE    = "OUTPUT_RATE"

    def name(self):
        return "beregn_omsaetningsrate"

    def displayName(self):
        return "Beregn Omsætningsrate (Oversvømmelse)"

    def group(self):
        return "Oversvoemmelse"

    def groupId(self):
        return "oversvoemmelse"

    def createInstance(self):
        return BeregnOmsaetningsrate()

    def initAlgorithm(self, config=None):
        self.addParameter(
            QgsProcessingParameterVectorLayer(
                self.PROJEKTOMRAADE,
                "Projektområde",
            )
        )
        self.addOutput(
            QgsProcessingOutputNumber(
                self.OUTPUT_RATE,
                "Omsætningsrate (kg N/ha/døgn)",
            )
        )

    def processAlgorithm(self, parameters, context: QgsProcessingContext, feedback: QgsProcessingFeedback):
        try:
            import openpyxl
        except ImportError:
            raise QgsProcessingException(
                "openpyxl er ikke installeret. Installer via OSGeo4W Shell: pip install openpyxl"
            )

        lag = self.parameterAsVectorLayer(parameters, self.PROJEKTOMRAADE, context)
        if not lag:
            raise QgsProcessingException("Projektområde-laget kunne ikke indlæses.")

        lag_navn = lag.name()

        projekt_mappe = QgsProject.instance().homePath()
        if not projekt_mappe:
            raise QgsProcessingException(
                "QGIS-projektet er ikke gemt. Gem projektet og prøv igen."
            )

        # Regnearket ligger i <projektområde>/Resultater — samme opslag som de
        # øvrige regnearkstrin bruger. Den gamle sti under QGIS-projektmappen
        # prøves bagefter, så en ældre projektmappe stadig virker.
        excel_filnavn = get_resultat_excel_navn()
        excel_sti = find_regneark()
        if not excel_sti:
            gammel_sti = os.path.join(projekt_mappe, f"Outputfiler_{lag_navn}",
                                      f"Resultat_{lag_navn}.xlsx")
            if os.path.isfile(gammel_sti):
                excel_sti = gammel_sti
        if not excel_sti:
            raise QgsProcessingException(
                f'{excel_filnavn} blev ikke fundet.\n\n'
                'Kør "Projektområde" i interfacet først — det opretter Excel-filen.\n'
                'Tjek også at mappe, projektnavn og projektområde er udfyldt i boks 1.'
            )

        feedback.pushInfo(f"Læser Excel-fil: {excel_sti}")

        # data_only=True for at få beregnede formelværdier (ikke formlerne selv)
        wb_read = openpyxl.load_workbook(excel_sti, data_only=True)

        if EXCEL_ARK_TILFOERSEL not in wb_read.sheetnames:
            raise QgsProcessingException(
                f'Arket "{EXCEL_ARK_TILFOERSEL}" blev ikke fundet i {excel_filnavn}.'
            )

        tn_raa = wb_read[EXCEL_ARK_TILFOERSEL][TN_BELASTNING_CELLE].value
        if tn_raa is None:
            # Cellen er en formel. Pluginnet skriver tal ind, men regner ikke
            # formler — så den gemte værdi står tom, til Excel har haft fat i
            # filen. Prøv at få den regnet igennem frem for at give op.
            wb_read.close()
            feedback.pushInfo(
                f"{EXCEL_ARK_TILFOERSEL}!{TN_BELASTNING_CELLE} har ingen beregnet "
                "værdi — regnearkets formler regnes igennem.")
            if genberegn_i_excel(excel_sti, feedback):
                wb_read = openpyxl.load_workbook(excel_sti, data_only=True)
                tn_raa = wb_read[EXCEL_ARK_TILFOERSEL][TN_BELASTNING_CELLE].value
        if tn_raa is None:
            raise QgsProcessingException(
                f"{EXCEL_ARK_TILFOERSEL}!{TN_BELASTNING_CELLE} (N-tab fra oplandet) "
                "har ingen beregnet værdi.\n\n"
                "Cellen er en formel, og formler regnes først ud når regnearket "
                "åbnes i Excel — pluginnet skriver kun tal ind.\n\n"
                f"Åbn regnearket, gem det, og kør trinnet igen:\n{excel_sti}\n\n"
                "Er felterne for nedbør, sandjord og dyrket areal tomme, mangler "
                'trinnene før dette — kør "Vandopland" og "Direkte opland".'
            )

        try:
            tn_belastning = float(tn_raa)
        except (TypeError, ValueError):
            raise QgsProcessingException(
                f'Ugyldig værdi i {EXCEL_ARK_TILFOERSEL}!{TN_BELASTNING_CELLE}: "{tn_raa}"'
            )

        feedback.pushInfo(f"TN belastning ({EXCEL_ARK_TILFOERSEL}!{TN_BELASTNING_CELLE}): {tn_belastning} kg N/ha/år")

        # Empirisk formel (Overvågning af vådområder 2018-2021, s. 74):
        #   TN fjernelse (%) = 27,5 * exp(-6e-4 * x)
        #   x = TN belastning (kg N/ha/år)
        tn_fjernelse_pct = 27.5 * math.exp(-6e-4 * tn_belastning)
        feedback.pushInfo(f"TN fjernelse (%): {tn_fjernelse_pct:.4f}")

        # Omregn til kg N/ha/døgn
        omsaetningsrate = (tn_fjernelse_pct / 100.0 * tn_belastning) / 365.0
        feedback.pushInfo(f"Omsætningsrate: {round(omsaetningsrate, 6)} kg N/ha/døgn")

        # Skriv til Omsætning!D32
        wb_write = openpyxl.load_workbook(excel_sti)
        if EXCEL_ARK_OMSAETNING not in wb_write.sheetnames:
            raise QgsProcessingException(
                f'Arket "{EXCEL_ARK_OMSAETNING}" blev ikke fundet i {excel_filnavn}.'
            )

        wb_write[EXCEL_ARK_OMSAETNING][OMSAETNINGSRATE_CELLE] = round(omsaetningsrate, 6)
        wb_write.save(excel_sti)

        feedback.pushInfo(
            f"Omsætningsrate skrevet til {EXCEL_ARK_OMSAETNING}!{OMSAETNINGSRATE_CELLE}: "
            f"{round(omsaetningsrate, 6)} kg N/ha/døgn"
        )

        return {self.OUTPUT_RATE: omsaetningsrate}
