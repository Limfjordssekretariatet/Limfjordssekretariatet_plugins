import os
import openpyxl

from qgis.core import QgsProcessingAlgorithm, QgsProcessingParameterVectorLayer, QgsProcessingException, QgsProject
from qgis import processing

# Sørg for at scriptets egen mappe er på sys.path, så 'utils' altid kan importeres.
import os as _os, sys as _sys
try:
    _d = _os.path.dirname(_os.path.abspath(__file__))
    if _d and _d not in _sys.path:
        _sys.path.insert(0, _d)
except NameError:
    pass
from utils import find_resultater_mappe, get_resultat_excel_navn


# ─────────────────────────────────────────────
# HJÆLPEFUNKTIONER
# ─────────────────────────────────────────────

def get_resultat_regneark_navn():
    """Returnerer dynamisk navn for resultat-regnearket (Resultat_<omraade>.xlsx)."""
    return get_resultat_excel_navn()


def find_regneark():
    """Returnerer Resultater/Resultat_<omraade>.xlsx (robust over for flade
    mappe-skemaer) eller None hvis filen ikke findes."""
    resultater = find_resultater_mappe()
    if not resultater:
        return None
    sti = os.path.join(resultater, get_resultat_excel_navn())
    return sti if os.path.isfile(sti) else None


def udtræk_cellid(udtræk_dmi_lag):
    """
    Læser 'cellid'-kolonnen fra Udtræk_DMI_GRID-laget og fjerner præfikset '10km_'.
    Returnerer en liste af rensede cellid-værdier.
    """
    cellid_værdier = []
    for feature in udtræk_dmi_lag.getFeatures():
        rå_cellid = feature["cellId"]
        renset = str(rå_cellid).replace("10km_", "").strip()
        cellid_værdier.append(renset)
    return cellid_værdier


def find_nedbor_i_regneark(regneark_sti, cellid_liste):
    """
    Slår cellid op i DMI-arket under kolonnen 'DMI10_NY'
    og returnerer den tilsvarende 'nedbor_kor'-værdi.
    """
    wb = openpyxl.load_workbook(regneark_sti, data_only=True, read_only=True)
    try:
        dmi_ark = wb["DMI"]

        headers = [cell.value for cell in dmi_ark[1]]
        col_dmi10_ny = headers.index("DMI10_NY")
        col_nedbor   = headers.index("nedbor_kor")

        for row in dmi_ark.iter_rows(min_row=2, values_only=True):
            if row[col_dmi10_ny] in cellid_liste:
                return row[col_nedbor]

        raise ValueError(f"Ingen match fundet i DMI-arket for cellid: {cellid_liste}")
    finally:
        wb.close()


def udtræk_s_procent(s_procent_lag):
    """
    Læser 'S_procent'-kolonnen fra S_procent-laget og returnerer værdien fra første feature.
    """
    for feature in s_procent_lag.getFeatures():
        return feature["S_procent"]
    raise ValueError("S_procent-laget indeholder ingen features.")


def udtræk_areal_ha(direkte_opland_areal_lag):
    """
    Læser 'Areal_ha'-kolonnen fra Direkte_Opland_Areal-laget og returnerer værdien fra første feature.
    """
    for feature in direkte_opland_areal_lag.getFeatures():
        return feature["Areal_ha"]
    raise ValueError("Direkte_Opland_Areal-laget indeholder ingen features.")


def udtræk_omdrift_procent(omdrift_procent_lag):
    """
    Læser overlap-procenten fra 'Beregn vektoroverlap'-laget (feltet der ender på _pc).

    Feltnavnet dannes af NAVNET på det lag der blev lagt henover, og det navn giver
    QGIS selv til et midlertidigt lag efter sin sprogindstilling: 'Beregnet' på
    dansk, 'Calculated' på engelsk. Et fast opslag på "Beregnet_pc" virker derfor
    kun på en dansk QGIS og fejler med KeyError på alle andre. Feltet findes i
    stedet på sin endelse.
    """
    felter = [f.name() for f in omdrift_procent_lag.fields()]
    kandidater = [n for n in felter if n.lower().endswith("_pc")]
    if not kandidater:
        raise ValueError(
            "Overlap-laget har intet procent-felt (_pc). Fundne felter: "
            + ", ".join(felter))
    # Dansk navn foretrækkes hvis det er der, ellers det første _pc-felt.
    navn = "Beregnet_pc" if "Beregnet_pc" in kandidater else kandidater[0]
    for feature in omdrift_procent_lag.getFeatures():
        return feature[navn]
    raise ValueError("Omdrift_procent-laget indeholder ingen features.")


def skriv_til_regneark(regneark_sti, værdier: dict):
    """
    Skriver værdier til specifikke celler i arket '1. Tilførsel'.
    værdier er en dict på formen {celle: værdi}, fx {"C42": 968.12, "C44": 75.3}
    """
    wb = openpyxl.load_workbook(regneark_sti)
    try:
        ark = wb["1. Tilførsel"]
        for celle, værdi in værdier.items():
            ark[celle] = værdi
        wb.save(regneark_sti)
    finally:
        wb.close()


# ─────────────────────────────────────────────
# QGIS PROCESSING ALGORITHM
# ─────────────────────────────────────────────

class DirecteOplandScript(QgsProcessingAlgorithm):

    UDTRÆK_DMI_LAG          = "UDTRÆK_DMI_LAG"
    S_PROCENT_LAG           = "S_PROCENT_LAG"
    DIREKTE_OPLAND_AREAL_LAG = "DIREKTE_OPLAND_AREAL_LAG"
    OMDRIFT_PROCENT_LAG     = "OMDRIFT_PROCENT_LAG"

    def name(self):
        return "direkte_opland_script"

    def displayName(self):
        return "Direkte Opland Script"

    def group(self):
        return "Vaadomraade"

    def groupId(self):
        return "vaadomraade"

    def shortHelpString(self):
        return (
            "Skriver resultater fra QGIS-modellen til Regneark_Resultat.xlsx:\n"
            "  - Nedbør (nedbor_kor)        → '1. Tilførsel'!C42\n"
            "  - Sandjord (S_procent)       → '1. Tilførsel'!C44\n"
            "  - Omdriftsprocent            → '1. Tilførsel'!C46\n"
            "  - Direkte opland areal (ha)  → '1. Tilførsel'!C48"
        )

    def initAlgorithm(self, config=None):
        self.addParameter(
            QgsProcessingParameterVectorLayer(
                self.UDTRÆK_DMI_LAG,
                "Udtræk DMI GRID lag"
            )
        )
        self.addParameter(
            QgsProcessingParameterVectorLayer(
                self.S_PROCENT_LAG,
                "S_procent lag"
            )
        )
        self.addParameter(
            QgsProcessingParameterVectorLayer(
                self.DIREKTE_OPLAND_AREAL_LAG,
                "Direkte Opland Areal lag"
            )
        )
        self.addParameter(
            QgsProcessingParameterVectorLayer(
                self.OMDRIFT_PROCENT_LAG,
                "Omdrift procent lag"
            )
        )

    def processAlgorithm(self, parameters, context, feedback):
        udtræk_dmi_lag           = self.parameterAsVectorLayer(parameters, self.UDTRÆK_DMI_LAG, context)
        s_procent_lag            = self.parameterAsVectorLayer(parameters, self.S_PROCENT_LAG, context)
        direkte_opland_areal_lag = self.parameterAsVectorLayer(parameters, self.DIREKTE_OPLAND_AREAL_LAG, context)
        omdrift_procent_lag      = self.parameterAsVectorLayer(parameters, self.OMDRIFT_PROCENT_LAG, context)

        # Find regneark
        regneark_sti = find_regneark()
        if not regneark_sti:
            raise QgsProcessingException(
                f"Kunne ikke finde '{REGNEARK_NAVN}' i Resultater-mappen. "
                f"Kør først 'Udvælg oplande model' (eller 'Skriv Arealer til Excel') "
                f"så regnearket bliver kopieret til Resultater/."
            )
        feedback.pushInfo(f"Regneark fundet: {regneark_sti}")

        # Trin 1: Nedbør fra DMI
        cellid_liste = udtræk_cellid(udtræk_dmi_lag)
        feedback.pushInfo(f"Fundet cellid(s): {cellid_liste}")
        nedbor = find_nedbor_i_regneark(regneark_sti, cellid_liste)
        feedback.pushInfo(f"Nedbør: {nedbor:.2f} mm → C42")

        # Trin 2: Sandjordsprocent
        s_procent = udtræk_s_procent(s_procent_lag)
        feedback.pushInfo(f"S_procent: {s_procent:.2f} % → C44")

        # Trin 3: Omdriftsprocent
        omdrift_procent = udtræk_omdrift_procent(omdrift_procent_lag)
        feedback.pushInfo(f"Omdrift_procent: {omdrift_procent:.2f} % → C46")

        # Trin 4: Direkte opland areal (m² → ha)
        areal_ha = udtræk_areal_ha(direkte_opland_areal_lag)
        feedback.pushInfo(f"Areal: {areal_ha:.2f} ha → C48")

        # Skriv alle værdier til regnearket i én operation (maks 2 decimaler)
        skriv_til_regneark(regneark_sti, {
            "C42": round(nedbor, 2),
            "C44": round(s_procent, 2),
            "C46": round(omdrift_procent, 2),
            "C48": round(areal_ha, 2),
        })
        feedback.pushInfo("Alle værdier skrevet til '1. Tilførsel'.")

        return {}

    def createInstance(self):
        return DirecteOplandScript()
