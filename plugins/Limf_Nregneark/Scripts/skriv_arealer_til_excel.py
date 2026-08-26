"""
Skriv Arealer til Excel
-----------------------
Modtager de fem beregnede areal-lag direkte fra feltberegnerne
i modellen og skriver totalarealerne (i hektar) til en kopi af
Excel-filen i mappen Resultat_regneark.

Kobl scriptets inputs til modeloutputs saaledes:
  Areal af agerjord  -> Agerjord-laget  (felt: Areal)
  Areal af brak      -> Brak-laget      (felt: Areal)
  Areal af graes     -> Graes-laget     (felt: Areal)
  Areal af natur     -> natur_area-laget (felt: Areal)
  Areal af befaestet -> Befaestet-laget (felt: Areal)

Excel-kildefil hentes automatisk fra:
  <rod>/Grunddata/mst_n_beregning_jul2023.xlsx
Output gemmes i:
  <rod>/<projektnavn>/<omraade>/Resultater/Regneark_Resultat.xlsx
"""

import os
import shutil

from qgis.core import (
    QgsProcessingAlgorithm,
    QgsProcessingParameterVectorLayer,
    QgsProcessingException,
    QgsVectorLayer,
    QgsProcessing,
)

GRUNDDATA_REGNEARK = "mst_n_beregning_jul2023.xlsx"


def get_resultat_regneark_navn():
    """Returnerer dynamisk navn for resultat-regnearket baseret på projektomraade-navn."""
    from qgis.core import QgsSettings
    s = QgsSettings()
    omraade = s.value('vaadomraade_modeller/projektomraade_navn', '')
    if not omraade:
        return "Resultat.xlsx"  # Fallback hvis omraade ikke er tilgængelig
    return f"Resultat_{omraade}.xlsx"


# Grunddata følger altid med pluginnet — hentes centralt fra utils.
# Sørg for at scriptets egen mappe er på sys.path, så 'utils' altid kan
# importeres uanset hvordan QGIS' script-provider loader filen.
import os as _os, sys as _sys
try:
    _d = _os.path.dirname(_os.path.abspath(__file__))
    if _d and _d not in _sys.path:
        _sys.path.insert(0, _d)
except NameError:
    pass  # __file__ ikke defineret i denne load-kontekst
from utils import find_grunddata, find_resultater_mappe  # noqa: F401


def find_resultater():
    """Returnerer Resultater/ (robust over for flade mappe-skemaer). Dette script
    SKRIVER til mappen, så den oprettes hvis projektmappen findes."""
    return find_resultater_mappe()

M2_TIL_HEKTAR = 10_000
EXCEL_ARK     = "1. Tilf\u00f8rsel"
EXCEL_KOLONNE = 3


class SkrivArealerTilExcel(QgsProcessingAlgorithm):

    AGERJORD  = "AGERJORD"
    BRAK      = "BRAK"
    GRAES     = "GRAES"
    NATUR     = "NATUR"
    BEFAESTET = "BEFAESTET"

    def name(self):
        return "skriv_arealer_til_excel"

    def displayName(self):
        return "Skriv Arealer til Excel"

    def group(self):
        return "Projektomraade"

    def groupId(self):
        return "projektomraade"

    def createInstance(self):
        return SkrivArealerTilExcel()

    def initAlgorithm(self, config=None):
        self.addParameter(
            QgsProcessingParameterVectorLayer(
                self.AGERJORD,
                "Areal af agerjord",
                types=[QgsProcessing.TypeVectorAnyGeometry],
            )
        )
        self.addParameter(
            QgsProcessingParameterVectorLayer(
                self.BRAK,
                "Areal af brak",
                types=[QgsProcessing.TypeVectorAnyGeometry],
            )
        )
        self.addParameter(
            QgsProcessingParameterVectorLayer(
                self.GRAES,
                "Areal af graes",
                types=[QgsProcessing.TypeVectorAnyGeometry],
            )
        )
        self.addParameter(
            QgsProcessingParameterVectorLayer(
                self.NATUR,
                "Areal af natur",
                types=[QgsProcessing.TypeVectorAnyGeometry],
            )
        )
        self.addParameter(
            QgsProcessingParameterVectorLayer(
                self.BEFAESTET,
                "Areal af befaestet",
                types=[QgsProcessing.TypeVectorAnyGeometry],
            )
        )

    def _sum_felt(self, lag, feltnavn, feedback):
        """Summerer et felt i alle features. Returnerer 0 ved tomt lag eller manglende felt."""
        if lag is None or not lag.isValid():
            return 0.0
        if lag.featureCount() == 0:
            feedback.pushWarning(f"Laget '{lag.name()}' er tomt – bruger 0.")
            return 0.0
        felter = [f.name() for f in lag.fields()]
        if feltnavn not in felter:
            feedback.pushWarning(
                f"Feltet '{feltnavn}' ikke fundet i '{lag.name()}' "
                f"(tilgaengelige: {felter}) – bruger 0."
            )
            return 0.0
        total = 0.0
        for feature in lag.getFeatures():
            val = feature[feltnavn]
            if val is not None:
                total += float(val)
        return total

    def _m2_til_ha(self, m2):
        return round(m2 / M2_TIL_HEKTAR, 4)

    def processAlgorithm(self, parameters, context, feedback):
        import openpyxl

        # Modtag de fem beregnede lag fra modellen
        agerjord  = self.parameterAsVectorLayer(parameters, self.AGERJORD,  context)
        brak      = self.parameterAsVectorLayer(parameters, self.BRAK,      context)
        graes     = self.parameterAsVectorLayer(parameters, self.GRAES,     context)
        natur     = self.parameterAsVectorLayer(parameters, self.NATUR,     context)
        befaestet = self.parameterAsVectorLayer(parameters, self.BEFAESTET, context)

        # Beregn totale arealer — alle lag bruger feltet "Areal"
        agerjord_ha  = self._m2_til_ha(self._sum_felt(agerjord,  "Areal", feedback))
        brak_ha      = self._m2_til_ha(self._sum_felt(brak,      "Areal", feedback))
        graes_ha     = self._m2_til_ha(self._sum_felt(graes,     "Areal", feedback))
        natur_ha     = self._m2_til_ha(self._sum_felt(natur,     "Areal", feedback))
        befaestet_ha = self._m2_til_ha(self._sum_felt(befaestet, "Areal", feedback))

        feedback.pushInfo(f"Agerjord:  {agerjord_ha} ha")
        feedback.pushInfo(f"Brak:      {brak_ha} ha")
        feedback.pushInfo(f"Graes:     {graes_ha} ha")
        feedback.pushInfo(f"Natur:     {natur_ha} ha")
        feedback.pushInfo(f"Befaestet: {befaestet_ha} ha")

        # Find Grunddata- og Resultater-mapperne via interfacet
        grunddata_mappe = find_grunddata()
        resultater_mappe = find_resultater()
        if not grunddata_mappe:
            raise QgsProcessingException(
                "Grunddata-mappen blev ikke fundet. "
                "Udpeg vedlagte mappe i interfacet først."
            )
        if not resultater_mappe:
            raise QgsProcessingException(
                "Resultater-mappen blev ikke fundet. "
                "Kør 'Navngiv projekt' og 'Tilføj dit projektområde (.shp)' i interfacet først."
            )

        kilde_excel  = os.path.join(grunddata_mappe, GRUNDDATA_REGNEARK)
        OUTPUT_EXCEL = os.path.join(resultater_mappe, get_resultat_regneark_navn())

        # Kopiér kildefilen til Resultater/ hvis den ikke allerede er der
        if not os.path.isfile(OUTPUT_EXCEL):
            if not os.path.isfile(kilde_excel):
                raise QgsProcessingException(
                    f"Excel-kildefilen blev ikke fundet: {kilde_excel}"
                )
            shutil.copy2(kilde_excel, OUTPUT_EXCEL)
            feedback.pushInfo(f"Kopi oprettet: {OUTPUT_EXCEL}")

        # Skriv arealer til kopien
        wb = openpyxl.load_workbook(OUTPUT_EXCEL)
        try:
            ws = wb[EXCEL_ARK] if EXCEL_ARK in wb.sheetnames else wb.active

            for raekke, navn, vaerdi in [
                (60, "Agerjord",    agerjord_ha),
                (61, "Ager, brak",  brak_ha),
                (62, "Vedv. graes", graes_ha),
                (63, "Natur*",      natur_ha),
                (64, "Befaestet",   befaestet_ha),
            ]:
                ws.cell(row=raekke, column=EXCEL_KOLONNE).value = vaerdi
                feedback.pushInfo(f"  C{raekke} ({navn}): {vaerdi} ha")

            # Middelværdier af N-udvaskning-intervaller (J60-62) → G60-62
            for raekke, navn, vaerdi in [
                (60, "Agerjord N-udvaskning",    47.5),  # interval 45-50 kg N/ha
                (61, "Ager brak N-udvaskning",    7.5),  # interval 5-10 kg N/ha
                (62, "Vedv. graes N-udvaskning",  2.5),  # interval 0-5 kg N/ha
            ]:
                ws.cell(row=raekke, column=7).value = vaerdi
                feedback.pushInfo(f"  G{raekke} ({navn}): {vaerdi} kg N/ha")

            wb.save(OUTPUT_EXCEL)
            feedback.pushInfo(f"Resultatfil gemt: {OUTPUT_EXCEL}")
        finally:
            wb.close()
        return {}
