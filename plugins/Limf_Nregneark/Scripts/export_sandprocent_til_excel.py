"""
Export Sandprocent til Excel
----------------------------
Læser S_procent værdien fra Jordtypemodellens outputlag
og skriver den til celle C44 i arket "1. Tilførsel".

Læser desuden nedbor_kor fra det udvalgte DMI_GRID lag
ved at slå celleid op i arket "DMI" kolonne DMI10_NY
og skriver værdien til celle C42 i arket "1. Tilførsel".
"""

from qgis.core import (
    QgsProcessingAlgorithm,
    QgsProcessingParameterVectorLayer,
    QgsProcessingParameterFile,
    QgsProcessingException,
)
import os

ARK_TILFORSEL = "1. Tilførsel"
ARK_DMI       = "DMI"
CELLE_S       = "C44"
CELLE_NEDBOR  = "C42"
FELT_S        = "S_procent"
FELT_CELLEID  = "cellId"
KOL_DMI10_NY  = 0  # kolonne A
KOL_NEDBOR    = 2  # kolonne C


class ExportSandprocentTilExcel(QgsProcessingAlgorithm):

    INPUT      = "INPUT"
    DMI_LAYER  = "DMI_LAYER"
    EXCEL_FILE = "EXCEL_FILE"

    def name(self):
        return "export_sandprocent_til_excel"

    def displayName(self):
        return "Export Sandprocent til Excel"

    def group(self):
        return "Jordtypemodel"

    def groupId(self):
        return "jordtypemodel"

    def createInstance(self):
        return ExportSandprocentTilExcel()

    def initAlgorithm(self, config=None):  # noqa: ARG002
        self.addParameter(
            QgsProcessingParameterVectorLayer(
                self.INPUT,
                "S_procent outputlag (fra Jordtypemodel)"
            )
        )
        self.addParameter(
            QgsProcessingParameterVectorLayer(
                self.DMI_LAYER,
                "Udvalgt DMI_GRID lag"
            )
        )
        self.addParameter(
            QgsProcessingParameterFile(
                self.EXCEL_FILE,
                "Excel-fil (.xlsx)",
                extension="xlsx"
            )
        )

    def processAlgorithm(self, parameters, context, feedback):
        lag       = self.parameterAsVectorLayer(parameters, self.INPUT, context)
        dmi_lag   = self.parameterAsVectorLayer(parameters, self.DMI_LAYER, context)
        excel_sti = self.parameterAsFile(parameters, self.EXCEL_FILE, context)

        if lag is None:
            raise QgsProcessingException("Kunne ikke indlæse S_procent laget.")
        if dmi_lag is None:
            raise QgsProcessingException("Kunne ikke indlæse DMI_GRID laget.")

        # --- Hent S_procent fra første feature ---
        s_procent_vaerdi = None
        for feature in lag.getFeatures():
            felt_index = lag.fields().indexOf(FELT_S)
            if felt_index == -1:
                raise QgsProcessingException(
                    f"Feltet '{FELT_S}' blev ikke fundet i laget. "
                    f"Tilgængelige felter: {[f.name() for f in lag.fields()]}"
                )
            s_procent_vaerdi = feature[felt_index]
            break

        if s_procent_vaerdi is None:
            raise QgsProcessingException("S_procent laget indeholder ingen features.")

        feedback.pushInfo(f"S_procent værdi: {s_procent_vaerdi:.2f}%")

        # --- Hent celleid fra DMI_GRID og strip "10km_"-præfiks ---
        celleid = None
        for feature in dmi_lag.getFeatures():
            felt_index = dmi_lag.fields().indexOf(FELT_CELLEID)
            if felt_index == -1:
                raise QgsProcessingException(
                    f"Feltet '{FELT_CELLEID}' blev ikke fundet i DMI_GRID laget. "
                    f"Tilgængelige felter: {[f.name() for f in dmi_lag.fields()]}"
                )
            celleid = str(feature[felt_index]).replace("10km_", "")
            break

        if celleid is None:
            raise QgsProcessingException("DMI_GRID laget indeholder ingen features.")

        feedback.pushInfo(f"DMI celleid (renset): {celleid}")

        # --- Åbn Excel ---
        try:
            import openpyxl
        except ImportError:
            raise QgsProcessingException(
                "Pakken 'openpyxl' er ikke installeret. Kør følgende i QGIS Python Console:\n"
                "import subprocess, sys; subprocess.run([sys.executable, '-m', 'pip', 'install', 'openpyxl'])"
            )

        if not os.path.exists(excel_sti):
            raise QgsProcessingException(f"Excel-filen blev ikke fundet: {excel_sti}")

        wb = openpyxl.load_workbook(excel_sti)

        # --- Slå celleid op i DMI-arket og hent nedbor_kor ---
        if ARK_DMI not in wb.sheetnames:
            raise QgsProcessingException(
                f"Arket '{ARK_DMI}' blev ikke fundet. "
                f"Tilgængelige ark: {wb.sheetnames}"
            )

        ws_dmi = wb[ARK_DMI]
        nedbor_kor = None
        for row in ws_dmi.iter_rows(min_row=2, values_only=True):
            if str(row[KOL_DMI10_NY]) == celleid:
                nedbor_kor = row[KOL_NEDBOR]
                break

        if nedbor_kor is None:
            raise QgsProcessingException(
                f"DMI celleid '{celleid}' blev ikke fundet i kolonnen DMI10_NY. "
                f"Tjek at celleid-formatet matcher (f.eks. '627_44')."
            )

        feedback.pushInfo(f"nedbor_kor: {nedbor_kor}")

        # --- Skriv begge værdier til "1. Tilførsel" og gem ---
        if ARK_TILFORSEL not in wb.sheetnames:
            raise QgsProcessingException(
                f"Arket '{ARK_TILFORSEL}' blev ikke fundet. "
                f"Tilgængelige ark: {wb.sheetnames}"
            )

        ws = wb[ARK_TILFORSEL]
        ws[CELLE_NEDBOR] = round(float(nedbor_kor), 2)
        ws[CELLE_S]      = round(s_procent_vaerdi, 2)
        wb.save(excel_sti)

        feedback.pushInfo(f"Skrev nedbor_kor={nedbor_kor} til {ARK_TILFORSEL} > {CELLE_NEDBOR}")
        feedback.pushInfo(f"Skrev S_procent={s_procent_vaerdi:.2f} til {ARK_TILFORSEL} > {CELLE_S}")

        return {}
