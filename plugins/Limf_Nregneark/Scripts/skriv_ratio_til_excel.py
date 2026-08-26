import os
from qgis.core import (
    QgsProcessingAlgorithm,
    QgsProcessingParameterVectorLayer,
    QgsSettings,
)
# Sørg for at scriptets egen mappe er på sys.path, så 'utils' altid kan
# importeres uanset hvordan QGIS' script-provider loader filen.
import os as _os, sys as _sys
try:
    _d = _os.path.dirname(_os.path.abspath(__file__))
    if _d and _d not in _sys.path:
        _sys.path.insert(0, _d)
except NameError:
    pass  # __file__ ikke defineret i denne load-kontekst
from utils import find_resultater_mappe, get_resultat_excel_navn  # noqa: F401

RATIO_FELT = 'ratio_pct'
ARK_NAVN = '2. Omsætning'
CELLE = 'D44'


def find_excel_fil(feedback):
    """Finder <rod>/<projektnavn>/<omraade>/Resultater/Resultat_<omraade>.xlsx."""
    resultater = find_resultater_mappe()
    if not resultater:
        feedback.reportError(
            'Resultater-mappen blev ikke fundet. '
            'Kør "Navngiv projekt" og "Tilføj dit projektområde (.shp)" i interfacet først.'
        )
        return None

    omraade = QgsSettings().value('vaadomraade_modeller/projektomraade_navn', '')
    excel_filnavn = f'Resultat_{omraade}.xlsx'
    excel_sti = os.path.join(resultater, excel_filnavn)

    feedback.pushInfo(f'Leder efter Excel-fil: {excel_sti}')

    if not os.path.isfile(excel_sti):
        feedback.reportError(
            f'Kunne ikke finde "{excel_filnavn}" i:\n{resultater}'
        )
        return None

    return excel_sti


class SkrivRatioTilExcel(QgsProcessingAlgorithm):
    INPUT = 'INPUT'

    def initAlgorithm(self, config=None):
        self.addParameter(
            QgsProcessingParameterVectorLayer(self.INPUT, 'Lag med ratio')
        )

    def processAlgorithm(self, parameters, context, feedback):
        layer = self.parameterAsVectorLayer(parameters, self.INPUT, context)

        # Hent ratio-værdien
        feltnavne = [f.name() for f in layer.fields()]
        feedback.pushInfo(f'Felter i laget: {", ".join(feltnavne)}')

        if RATIO_FELT not in feltnavne:
            feedback.reportError(
                f'Feltet "{RATIO_FELT}" blev ikke fundet i laget.\n'
                f'Tilgængelige felter: {", ".join(feltnavne)}\n'
                f'Ret RATIO_FELT øverst i scriptet til det korrekte feltnavn.'
            )
            return {}

        ratio = None
        for feature in layer.getFeatures():
            ratio = feature[RATIO_FELT]
            break

        if ratio is None:
            feltnavne = [f.name() for f in layer.fields()]
            feedback.reportError(
                f'Feltet "{RATIO_FELT}" blev ikke fundet i laget.\n'
                f'Tilgængelige felter: {", ".join(feltnavne)}'
            )
            return {}

        feedback.pushInfo(f'Ratio aflæst: {ratio}')

        # Beregn kvælstoffjernelseseffektivitet: y = 22,9 + 11,8 * ln(x)
        # ratio er allerede i procent (ratio_pct fra feltberegneren)
        import math
        x = ratio
        if x <= 0:
            feedback.reportError(f'Ratio er {x:.2f}% — kan ikke beregne ln af 0 eller negativt tal.')
            return {}
        value = round(22.9 + 11.8 * math.log(x), 2)
        feedback.pushInfo(f'y = 22,9 + 11,8 * ln({x:.2f}) = {value}')
        if value < 0:
            feedback.reportError(f'Resultat er negativt ({value}) — tjek om ratioen er realistisk.')
            return {}
        if value > 100:
            feedback.reportError(f'Resultat er over 100% ({value}) — tjek om ratioen er realistisk.')
            return {}

        # Find Excel-filen automatisk
        excel_sti = find_excel_fil(feedback)
        if excel_sti is None:
            return {}

        # Skriv til Excel
        try:
            import openpyxl
        except ImportError:
            feedback.reportError(
                'Pakken "openpyxl" er ikke installeret.\n'
                'Åbn OSGeo4W Shell og kør: python -m pip install openpyxl'
            )
            return {}

        wb = openpyxl.load_workbook(excel_sti)

        if ARK_NAVN not in wb.sheetnames:
            feedback.reportError(
                f'Arket "{ARK_NAVN}" blev ikke fundet.\n'
                f'Tilgængelige ark: {", ".join(wb.sheetnames)}'
            )
            return {}

        wb[ARK_NAVN][CELLE] = value
        wb.save(excel_sti)

        feedback.pushInfo(f'Gemt: {excel_sti}')
        return {}

    def name(self):
        return 'skriv_ratio_til_excel'

    def displayName(self):
        return 'Skriv ratio til Excel'

    def group(self):
        return 'Omsætning'

    def groupId(self):
        return 'omsaetning'

    def shortHelpString(self):
        return (
            f'Læser "{RATIO_FELT}" fra ratio-laget og beregner '
            f'y = 22,9 + 11,8 * ln(ratio). Skriver resultatet til celle {CELLE} '
            f'i arket "{ARK_NAVN}" i Resultat_<omraade>.xlsx.\n\n'
            f'Excel-filen søges automatisk frem i Resultater-mappen.'
        )

    def createInstance(self):
        return SkrivRatioTilExcel()
